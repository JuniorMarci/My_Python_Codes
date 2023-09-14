def remNA(caminho):
    from openpyxl import load_workbook

    filename = caminho
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'N/A':
                cell.value = -1

    workbook.save(caminho)


import pandas as pd
from datetime import datetime

#Data = ["Dados atualizados em: "+input("Dados atualizados em: ")]
Data = ["Dados atualizados em: "+str(datetime.now().strftime("%d/%m/%y às %H:%M:%S"))]


# Carregando a tabela de informações consolidadas
ConsPres = './Entradas_Mereo/Consolidada_Presidencia.xlsx'
ConsEmp = './Entradas_Mereo/Consolidada_Empreendimentos.xlsx'
remNA(ConsPres)
remNA(ConsEmp)
ConsPres = pd.read_excel(ConsPres)
ConsEmp = pd.read_excel(ConsEmp)

Consolidada = pd.concat([ConsPres, ConsEmp], axis=0, ignore_index=True)
Consolidada['Cod.Unico'] ="C"+Consolidada['Código da Área']+"::"+Consolidada['Login']

# Carregando a tabela de informações separados por metas
MetPres = './Entradas_Mereo/Meta_Presidencia.xlsx'
MetEmp = './Entradas_Mereo/Meta_Empreendimentos.xlsx'
remNA(MetPres)
remNA(MetEmp)
MetPres = pd.read_excel(MetPres)
MetEmp = pd.read_excel(MetEmp)

Metas = pd.concat([MetPres, MetEmp], axis=0, ignore_index=True)
Metas['Cod.Unico'] ="C"+Metas['Código da Área']+"::"+Metas['Login']

# Carregando a tabela de informações dos shoppings
sh = pd.read_excel('./Entradas_Personalizadas/Shoppings.xlsx')

tabCat = [] # Criada variável da nova tabela de categorias
tabCons = [] # Criada variável dos dados consolidados separados por mês


def Separar(texto):
    nt = []
    sep = [' ','/','(',')']
    for c in sep:
        texto = texto.replace(c,'-')
    return texto

for c in range(len(Consolidada['Cod.Unico'])): # Identificação dos códigos únicos
    #df['sh'][c] = [] # Criação da base da coluna
    idc = []

    for d in (Separar(Consolidada['Descrição da Área'][c].replace("Partage ADM","Partage_ADM")).split('-')): # Iteração nas áreas já formatadas
        if (d in sh['CS'].values): # Comparação com cara código de shopping
            #df['sh'][c].append(d) # Acréscimo na coluna. Tentativa de fazer tudo numa só tabela
            tabCat.append([Consolidada['Cod.Unico'][c],d]) # Alimentação da tabela de categorias
            idc.append(c)

    if (c in idc): 1==1
    else: tabCat.append([Consolidada['Cod.Unico'][c],'Sem categoria'])

    for m in range(1,9):
        actual = "mês "+str(m)
        tabCons.append([Consolidada['Cod.Unico'][c],m,Consolidada[actual][c]])


DFCat = pd.DataFrame(tabCat,columns=['Cod.Unico','Categorias']) # Junção inteligente das tabelas, a fim de criar uma relação das metas com as categorias

'''
MetUn = pd.DataFrame(Metas['Código da Meta'].unique(),columns=['CM'])
ShMet = []
for d in MetUn:
    ShMet.append(Metas[Metas['Código da Meta'] == d]['Cod.Unico'].iloc[0])

ShMet = pd.DataFrame({'Meta':MetUn,'Shopping':ShMet})#.to_csv('Saída_Algoritmo/ShoppingMeta.csv')
'''


pd.DataFrame(tabCons).to_csv('Saída_Algoritmo/DConsolidada.csv',sep=';',index=False)
pd.DataFrame(tabCat).to_csv('Saída_Algoritmo/Categorias.csv',sep=';',index=False)
Consolidada.to_csv('Saída_Algoritmo/Consolidada.csv',sep=';')
Metas.to_csv('Saída_Algoritmo/Metas.csv',sep=';')
DFCat.join(Metas[['Cod.Unico','Código da Meta']].set_index('Cod.Unico'),on='Cod.Unico').drop_duplicates().to_csv('Saída_Algoritmo/Metas_com_Categorias.csv',sep=';')
#df.to_csv('Dados.csv',sep=';')

#----------------------------------------------

# Carregando a tabela de Planos de Ações
PlA = './Entradas_Mereo/Planos de Ação/Plano.xlsx'
PlA = pd.read_excel(PlA)
Planos = [] # Criada tabela para correlacionar os meses com os planos
for c in range (len(PlA['Código da Ação'])):
    for d in range (int(PlA['Data início planejada'][c][3:5]),int(PlA['Data fim planejada'][c][3:5])+1):
        Planos.append([PlA['Código da Ação'][c],d])

pd.DataFrame(Planos).to_csv('./Saída_Algoritmo/Planos-Mês.csv',sep=';',index=False)

pd.DataFrame(Data).to_csv('./Saída_Algoritmo/Data de Atualizacao.txt',index=False,header=None)

print("Fim!")
